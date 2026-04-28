#!/usr/bin/env python3
"""Compare old vs new March 2026 reports — fail on any non-terminology diff."""
import os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

BASE = r"C:\Users\דרורנדל\HASHOMER HACHADASH\כספים - General\כספים - תכנון ובקרה\אדם ואדמה\2026\בקרות חודשיות\הכנסות מחקלאים\ניתוח פערים\דוחות"
OLD_DIR = os.path.join(BASE, "2026-03-מרץ_v2")
NEW_DIR = os.path.join(BASE, "2026-03-מרץ")

PAIRS = [
    ("אשכול_2026.xlsx",       "ניתוח פערים אשכול 03-26.xlsx"),
    ("גולן_2026.xlsx",        "ניתוח פערים גולן 03-26.xlsx"),
    ("גליל עליון_2026.xlsx",  "ניתוח פערים גליל עליון 03-26.xlsx"),
    ("חצבה_2026.xlsx",        "ניתוח פערים חצבה 03-26.xlsx"),
    ("ניצנה_2026.xlsx",       "ניתוח פערים ניצנה 03-26.xlsx"),
    ("ניצנים_2026.xlsx",      "ניתוח פערים ניצנים 03-26.xlsx"),
    ("תבור_2026.xlsx",        "ניתוח פערים תבור 03-26.xlsx"),
    ("סיכום_רשת_2026_03.xlsx", "ניתוח פערים סיכום רשת 03-26.xlsx"),
]

def normalize(v):
    """Treat חניך/עובד as equal so terminology change is allowed."""
    if isinstance(v, str):
        return v.replace("חניך", "עובד")
    return v

def cmp_workbook(old_path, new_path):
    wb_o = load_workbook(old_path, data_only=False)
    wb_n = load_workbook(new_path, data_only=False)
    diffs = []

    # Map normalized sheet name → real name in each workbook
    norm_o = {normalize(s): s for s in wb_o.sheetnames}
    norm_n = {normalize(s): s for s in wb_n.sheetnames}
    only_old = set(norm_o) - set(norm_n)
    only_new = set(norm_n) - set(norm_o)
    if only_old:
        diffs.append(f"  ⓘ לשוניות רק בישן: {[norm_o[k] for k in only_old]}")
    if only_new:
        diffs.append(f"  ⓘ לשוניות רק בחדש: {[norm_n[k] for k in only_new]}")

    common = sorted(set(norm_o) & set(norm_n))
    for ns in common:
        sn_o, sn_n = norm_o[ns], norm_n[ns]
        ws_o, ws_n = wb_o[sn_o], wb_n[sn_n]
        sn = sn_o  # for display
        if (ws_o.max_row, ws_o.max_column) != (ws_n.max_row, ws_n.max_column):
            diffs.append(f"  [{sn}] dims old={ws_o.max_row}x{ws_o.max_column} "
                         f"new={ws_n.max_row}x{ws_n.max_column}")
        rows = max(ws_o.max_row, ws_n.max_row)
        cols = max(ws_o.max_column, ws_n.max_column)
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                vo = ws_o.cell(r, c).value
                vn = ws_n.cell(r, c).value
                if normalize(vo) != normalize(vn):
                    diffs.append(f"  [{sn}] R{r}C{c}: old={vo!r} | new={vn!r}")
                    if len(diffs) > 50:
                        diffs.append("  ... (truncated)")
                        return diffs
    return diffs

total = 0
for old_name, new_name in PAIRS:
    op = os.path.join(OLD_DIR, old_name)
    np = os.path.join(NEW_DIR, new_name)
    print(f"\n=== {old_name}  ↔  {new_name} ===")
    if not os.path.exists(op):
        print(f"  MISSING OLD: {op}"); continue
    if not os.path.exists(np):
        print(f"  MISSING NEW: {np}"); continue
    diffs = cmp_workbook(op, np)
    if not diffs:
        print("  ✅ זהה (פרט להחלפת חניך→עובד)")
    else:
        print(f"  ⚠ נמצאו {len(diffs)} הבדלים:")
        for d in diffs:
            print(d)
        total += len(diffs)

# Compare warnings logs
old_log = os.path.join(OLD_DIR, "לוג_בעיות_2026_03.txt")
new_log = os.path.join(NEW_DIR, "לוג בעיות 03-26.txt")
print(f"\n=== לוג בעיות ===")
if os.path.exists(old_log) and os.path.exists(new_log):
    o = open(old_log, encoding="utf-8").read().replace("חניך", "עובד")
    n = open(new_log, encoding="utf-8").read().replace("חניך", "עובד")
    if o == n:
        print("  ✅ זהה")
    else:
        print("  ⚠ הבדל בלוג")
        # show first few differing lines
        ol = o.splitlines(); nl = n.splitlines()
        for i, (a, b) in enumerate(zip(ol, nl)):
            if a != b:
                print(f"  L{i+1}: old={a!r} | new={b!r}")
                total += 1
                if total > 60: break
        if len(ol) != len(nl):
            print(f"  line count: old={len(ol)} new={len(nl)}")

print(f"\n{'='*60}")
print(f"סה\"כ הבדלים שאינם של מינוח: {total}")
