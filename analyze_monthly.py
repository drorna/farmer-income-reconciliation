#!/usr/bin/env python3
"""
ניתוח חודשי – הכנסות מחקלאים
Monthly + YTD agricultural income gap analysis for all schools.

-------------------------------------------------------------------------------
תכונות:
  • תמיכה בשורות תשלום שאינן שעתיות — עמודה E בקובצי המרכז.
  • שני סוגים מזוהים לפי טקסט בעמודה E:
      – המכיל "תפוקה"  → סוג "תפוקה" (תשלום לפי פלט: דולבים, ק"ג וכו')
      – המכיל "יום" / "יומי" → סוג "יומי" (תשלום לפי אדם ליום)
  • שורות לא-שעתיות:
      – לא נכנסות לאף חישוב KPI שעתי (תעריף, שעות/עובד, עובדים/יום וכו')
      – לא מעוותות את ההכנסה הראשית (נשמרת טהורה — רק שעתי)
      – מוצגות בטבלת "הכנסה שאינה שעתית" בכל גיליון חודש + YTD של בית הספר
        עם עמודה "סוג" שמציגה "יומי" או "תפוקה"
      – מוצגות בלשונית "הכנסה שאינה שעתית" בחוברת סיכום הרשת
  • שורה לא-שעתית לא צריכה שעות (עמודה C), ולא נכנסת ללוג בעיות.
  • שורה לא-שעתית צריכה: תאריך (A) + תשלום (D) + הסימון (E).
  • מספר עובדים (B) — אופציונלי. בדר"כ יש לשורות יומיות, חסר בתפוקה. אם חסר → "אין נתון".

חישובים:
  • YTD של ביצוע — פריט-פריט על פני כל הימים של כל החודשים.
  • YTD של תקציב — משוקלל לפי ימי עבודה של כל חודש.
  • שורת "סה"כ רשת" — משוקללת לפי ימי עבודה של כל בית ספר (כל יום שווה).
  • לוח "תצוגת על" — לשונית ראשונה בחוברת סיכום הרשת, כל ה-KPIs של כל בתי הספר.
-------------------------------------------------------------------------------

Usage:
    python analyze_monthly.py [YYYY] [MM]
    (no args → current month)

Expected source format per school sheet (after monthly cleanup):
    Row 1 : headers  (A=תאריך, B=עובדים, C=שעות, D=תשלום, E=אופן תשלום [אופציונלי])
    Row 2+: one row per farmer×date  (date rows only, no farmer subtotals)
    Date can be: "DD-MMM"  /  datetime object  /  "YYYY-MM-DD HH:MM:SS"
    Column E (optional):
        – מכיל "תפוקה" → תשלום לפי תפוקה   (לדוגמה: "לפי תפוקה", "תפוקה - דולב 130")
        – מכיל "יום"   → תשלום יומי         (לדוגמה: "לפי יום", "יומי 180/אדם")
        – ריק/אחר      → שורה שעתית רגילה

Output folder:  <BASE_DIR>/דוחות/YYYY-MM-Month/
    • <School>_YYYY.xlsx   – per-school annual workbook
    • סיכום_רשת_YYYY_MM.xlsx – network summary
"""

import os, sys, io, re, warnings
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR    = r"C:\Users\דרורנדל\HASHOMER HACHADASH\כספים - General\כספים - תכנון ובקרה\אדם ואדמה\2026\בקרות חודשיות\הכנסות מחקלאים\ניתוח פערים"
DATA_DIR    = os.path.join(BASE_DIR, "נתונים מרוכזים")
OUT_DIR     = os.path.join(BASE_DIR, "דוחות")
BUDGET_FILE = os.path.join(DATA_DIR, "תקציב.xlsx")

# ── Constants ──────────────────────────────────────────────────────────────────
MONTH_HE = {1:"ינואר",2:"פברואר",3:"מרץ",4:"אפריל",5:"מאי",6:"יוני",
            7:"יולי",8:"אוגוסט",9:"ספטמבר",10:"אוקטובר",11:"נובמבר",12:"דצמבר"}
ABBR_HE  = {1:"ינו",2:"פבר",3:"מרץ",4:"אפר",5:"מאי",6:"יונ",
            7:"יול",8:"אוג",9:"ספט",10:"אוק",11:"נוב",12:"דצמ"}

# Sheets to skip in the monthly workbook (not school data)
SKIP_SHEETS = {"פרטי חקלאים"}
SKIP_PREFIX = "מרכז חשבוניות"

# Non-hourly type detection — column E substring matching.
# Order matters: "תפוקה" is checked first, so "יום תפוקה" → "תפוקה".
# Each entry: (list_of_substrings_to_match, type_label)
PIECE_TYPES = [
    (["תפוקה"],        "תפוקה"),
    (["יומי", "יום"],  "יומי"),
]

def _detect_piece_type(flag_raw):
    """
    Detect non-hourly payment type from column E value.
    Returns type label ("תפוקה" / "יומי") or None (= hourly, default).
    """
    if not isinstance(flag_raw, str):
        return None
    for markers, label in PIECE_TYPES:
        if any(m in flag_raw for m in markers):
            return label
    return None

# ── Style helpers ──────────────────────────────────────────────────────────────
FONT_NAME = "Arial"
C_TITLE = "1F4E79"; C_HDR = "2E74B5"
C_ACT   = "DEEAF1"; C_BUD = "E2EFDA"; C_GAP = "FCE4D6"; C_ALT = "EBF3FB"
C_WHITE = "FFFFFF"; C_BLACK = "000000"
C_GAP_POS = "C6EFCE"; C_GAP_POS_F = "375623"   # green  – actual > budget
C_GAP_NEG = "FFC7CE"; C_GAP_NEG_F = "9C0006"   # red    – actual < budget
C_PIECE   = "FFF2CC"; C_PIECE_HDR = "BF8F00"   # warm amber for piece-rate blocks

def _f(bold=False, size=10, color=C_BLACK):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color)
def _fill(c):  return PatternFill("solid", fgColor=c)
def _bdr(c="BFBFBF"):
    s = Side(style="thin", color=c); return Border(left=s,right=s,top=s,bottom=s)
def _al(h="center", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap, readingOrder=0)
def _s(cell, bold=False, size=10, fg=C_BLACK, bg=None, h="center", wrap=False, fmt=None):
    cell.font = _f(bold, size, fg); cell.alignment = _al(h, wrap)
    cell.border = _bdr()
    if bg:  cell.fill   = _fill(bg)
    if fmt: cell.number_format = fmt
def cw(ws, d):
    for k, v in d.items(): ws.column_dimensions[k].width = v

# ── Date helpers ───────────────────────────────────────────────────────────────
_DATE_RE = re.compile(r"^\d{1,2}-[א-ת]{2,4}$")   # "01-ינו"

def _is_date(v):
    """Return True if v looks like a work date (not a farmer name)."""
    if isinstance(v, datetime): return True
    if not isinstance(v, str):  return False
    v = v.strip()
    if _DATE_RE.match(v): return True
    try:                          # "YYYY-MM-DD ..." full datetime string
        datetime.fromisoformat(v[:10]); return True
    except Exception: return False

def _lbl(v):
    """Normalise any date value to 'DD-MMM' Hebrew label."""
    if isinstance(v, datetime):
        return f"{v.day:02d}-{ABBR_HE.get(v.month,'??')}"
    if isinstance(v, str):
        v = v.strip()
        if _DATE_RE.match(v): return v
        try:
            d = datetime.fromisoformat(v[:10])
            return f"{d.day:02d}-{ABBR_HE.get(d.month,'??')}"
        except Exception: pass
    return str(v)

def _flt(v):
    try:    return float(v) if v not in (None, "", " ") else 0.0
    except: return 0.0

def _piece_sort_key(row):
    """Sort key for non-hourly rows: (month, day, type). Labels like '06-ינו' → (1,6,type)."""
    lbl = str(row.get("date_label", ""))
    typ = str(row.get("piece_type", ""))
    if "-" in lbl:
        parts = lbl.split("-", 1)
        try:
            day = int(parts[0])
            month_abbr = parts[1].strip()
            for m, abbr in ABBR_HE.items():
                if abbr == month_abbr:
                    return (m, day, typ)
        except Exception:
            pass
    return (99, 99, typ)

# ── Source file path ───────────────────────────────────────────────────────────
def monthly_path(year, month):
    return os.path.join(DATA_DIR,
        f"מרכז חשבוניות חקלאים {MONTH_HE[month]} {str(year)[-2:]}.xlsx")

def school_sheets(wb):
    return [s for s in wb.sheetnames
            if s not in SKIP_SHEETS and not s.startswith(SKIP_PREFIX)]

# ── Read one school sheet ──────────────────────────────────────────────────────
def _find_header_row(ws):
    """Return index of header row (1-based). Looks for 'תאריך'/'חקלאי' in first 5 rows."""
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if isinstance(cell.value, str) and \
               any(kw in cell.value for kw in ("תאריך", "חקלאי", "שם חקלאי")):
                return cell.row
    return 1   # fallback

def read_school_ws(ws, school_rate=35):
    """
    Read date rows from a normalised school sheet (cols A-E).
    Column E (row[4]): if contains "תפוקה"/"יום" → non-hourly row of corresponding type.

    Returns (hourly_rows, piece_rows, warnings):
        hourly_rows – list of dicts: {date_label, workers, hours, payment}
                      (השורות הרגילות – כולן שעתיות)
        piece_rows  – list of dicts: {date_label, piece_type, workers (or None), payment}
                      (שורות לא-שעתיות – ללא שעות, נתונים חלקיים מותרים)
                      piece_type is "תפוקה" or "יומי"
        warnings    – list of human-readable issue strings (רק שורות שעתיות)

    Skips rows where col-A is not a date (farmer subtotals etc.).
    For hourly rows: flags rows where hours (col C) is missing — rate cannot be computed.
    For non-hourly rows: missing hours is NORMAL — no warning. Missing workers → stored as None.
    """
    hdr         = _find_header_row(ws)
    hourly_rows = []
    piece_rows  = []
    warnings    = []
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        a = row[0] if len(row) > 0 else None
        if not _is_date(a):
            continue                        # farmer summary / header → skip

        lbl         = _lbl(a)
        workers_raw = row[1] if len(row) > 1 else None
        hours       = _flt(row[2] if len(row) > 2 else 0)
        payment     = _flt(row[3] if len(row) > 3 else 0)
        flag_raw    = row[4] if len(row) > 4 else None

        # Detect non-hourly payment type from column E
        piece_type = _detect_piece_type(flag_raw)

        if payment == 0:
            continue                        # truly empty row

        if piece_type is not None:
            # שורה לא-שעתית (יומי/תפוקה): לא דורשת שעות. עובדים אופציונלי.
            workers_val = None
            if workers_raw not in (None, "", " "):
                try:
                    w = float(workers_raw)
                    if w > 0:
                        workers_val = w
                except Exception:
                    workers_val = None
            piece_rows.append({"date_label": lbl,
                               "piece_type": piece_type,
                               "workers":    workers_val,
                               "payment":    payment})
            continue

        # שורה שעתית (התנהגות מקורית)
        workers = _flt(workers_raw)
        if hours == 0:
            warnings.append(
                f"  שורה {lbl}: תשלום={payment:.0f} ₪ אך שעות חסרות — "
                f"התעריף לא ניתן לחישוב, השורה לא נכללת בניתוח")
            continue                        # cannot compute rate → skip

        hourly_rows.append({"date_label": lbl,
                            "workers":    workers,
                            "hours":      hours,
                            "payment":    payment})
    return hourly_rows, piece_rows, warnings

def aggregate_by_date(rows):
    """
    Aggregate multiple farmer rows for the same date into one daily row.
    Returns list of dicts: {date_label, workers, hours, payment, rate, hours_per_worker}
    """
    from collections import defaultdict
    by_date = defaultdict(lambda: {"workers":0.0,"hours":0.0,"payment":0.0})
    for r in rows:
        by_date[r["date_label"]]["workers"] += r["workers"]
        by_date[r["date_label"]]["hours"]   += r["hours"]
        by_date[r["date_label"]]["payment"] += r["payment"]
    result = []
    for lbl, d in by_date.items():
        rate = d["payment"] / d["hours"] if d["hours"] else 0
        hpw  = d["hours"] / d["workers"] if d["workers"] else 0
        result.append({"date_label":      lbl,
                        "workers":         d["workers"],
                        "hours":           d["hours"],
                        "payment":         d["payment"],
                        "rate":            rate,
                        "hours_per_worker":hpw})
    return result

def aggregate_piece_by_date(rows):
    """
    Aggregate non-hourly rows by (date, piece_type).
    Two farmers on the same date with different types → kept as two separate rows.
    Workers: summed ONLY if all rows for that (date,type) report workers; otherwise None.
    Returns list of dicts: {date_label, piece_type, workers (or None), payment}
    """
    from collections import defaultdict
    by_key = defaultdict(lambda: {"workers":0.0,"workers_known":True,"payment":0.0})
    for r in rows:
        key = (r["date_label"], r["piece_type"])
        d = by_key[key]
        d["payment"] += r["payment"]
        if r["workers"] is None:
            d["workers_known"] = False
        else:
            d["workers"] += r["workers"]
    out = []
    for (lbl, typ), d in by_key.items():
        out.append({"date_label": lbl,
                    "piece_type": typ,
                    "workers":    d["workers"] if d["workers_known"] and d["workers"] > 0 else None,
                    "payment":    d["payment"]})
    return out

# ── Budget ─────────────────────────────────────────────────────────────────────
_bud_cache = {}

def _budget_wb():
    if "wb" not in _bud_cache:
        _bud_cache["wb"] = load_workbook(BUDGET_FILE, data_only=True) \
                           if os.path.exists(BUDGET_FILE) else None
    return _bud_cache["wb"]

def read_budget(school, month):
    """Return budget dict for school/month, or None if not found."""
    wb = _budget_wb()
    if wb is None or school not in wb.sheetnames:
        return None
    ws = wb[school]
    col = None
    for cell in ws[1]:
        if isinstance(cell.value, datetime) and cell.value.month == month:
            col = cell.column; break
    if col is None:
        return None
    def g(r): return _flt(ws.cell(row=r, column=col).value)
    students  = g(3); pct = g(4); staff = g(5)
    wpd       = students * pct + staff
    hpt       = g(7); rate = g(8)
    return {"work_days":         g(2),
            "workers_per_day":   wpd,
            "hours_per_day":     wpd * hpt,
            "hours_per_trainee": hpt,
            "total_revenue":     g(2) * wpd * hpt * rate,
            "rate":              rate}

def read_budget_rate(school):
    """Return the school's default hourly rate from budget (month 1), or 35."""
    b = read_budget(school, 1)
    return b["rate"] if b else 35.0

def ytd_budget(school, through_month):
    """
    YTD budget aggregation:
      Summable KPIs (work_days, total_revenue) → plain sum across months.
      Rate-style KPIs (workers_per_day, hours_per_day, hours_per_trainee, rate)
        → weighted average by work_days, so that each planned day contributes
          equally (consistent with how actual YTD works: day-by-day flattening).
    """
    months = [read_budget(school, m) for m in range(1, through_month + 1)]
    months = [m for m in months if m]
    if not months: return None
    total_days = sum(m["work_days"] for m in months)
    total_rev  = sum(m["total_revenue"] for m in months)
    if total_days == 0:
        # fallback: no planned work days → simple per-month averages
        n = len(months)
        return {"work_days":         0,
                "workers_per_day":   sum(m["workers_per_day"]   for m in months) / n,
                "hours_per_day":     sum(m["hours_per_day"]     for m in months) / n,
                "hours_per_trainee": sum(m["hours_per_trainee"] for m in months) / n,
                "total_revenue":     total_rev,
                "rate":              sum(m["rate"]              for m in months) / n}
    def w(key):   # weighted by work_days
        return sum(m["work_days"] * m[key] for m in months) / total_days
    return {"work_days":         total_days,
            "workers_per_day":   w("workers_per_day"),
            "hours_per_day":     w("hours_per_day"),
            "hours_per_trainee": w("hours_per_trainee"),
            "total_revenue":     total_rev,
            "rate":              w("rate")}

# ── KPI helpers ────────────────────────────────────────────────────────────────
# KPI order: [work_days, avg_workers, avg_hours_per_day, avg_hrs_per_worker, total_payment, rate]
KPI_HDRS = ["ימי עבודה", "ממוצע עובדים/יום", "ממוצע שעות/יום",
            "שעות/עובד", "הכנסה (₪)", "תעריף ממוצע בפועל (₪/ש)"]
KPI_FMTS = ["#,##0", "#,##0.0", "#,##0.0", "#,##0.0", "#,##0", "#,##0.00"]

def actuals_kpis(daily):
    """Compute 6 KPIs from a list of daily dicts."""
    if not daily:
        return [0, 0, 0, 0, 0, 0]
    n    = len(daily)
    tot_pay  = sum(d["payment"] for d in daily)
    tot_hrs  = sum(d["hours"]   for d in daily)
    return [n,
            sum(d["workers"]          for d in daily) / n,
            sum(d["hours"]            for d in daily) / n,
            sum(d["hours_per_worker"] for d in daily) / n,
            tot_pay,
            tot_pay / tot_hrs if tot_hrs else 0]   # weighted rate

def budget_kpis(bud):
    if bud is None: return None
    return [bud["work_days"], bud["workers_per_day"], bud["hours_per_day"],
            bud["hours_per_trainee"], bud["total_revenue"], bud["rate"]]

# ── Sheet writers ──────────────────────────────────────────────────────────────
DAILY_HDRS = ["תאריך", "עובדים", "סך שעות", "שעות/עובד", "הכנסה (₪)", "תעריף ממוצע בפועל (₪/ש)"]
DAILY_FMTS = ["@", "#,##0", "#,##0.0", "#,##0.0", "#,##0", "#,##0.00"]

def _summary_block(ws, act_r, act_vals, bud_vals):
    """Write ביצוע / תקציב / (blank) / פער / %פער rows starting at act_r."""
    bud_r = act_r + 1
    gap_r = act_r + 3
    pct_r = act_r + 4

    for label, rn, vals, bg in [("ביצוע", act_r, act_vals, C_ACT),
                                 ("תקציב", bud_r, bud_vals, C_BUD)]:
        _s(ws.cell(rn, 1, label), bold=True, bg=bg, h="right")
        for ci in range(2, 8):
            v = vals[ci-2] if vals else None
            _s(ws.cell(rn, ci, v), bold=True, bg=bg, fmt=KPI_FMTS[ci-2])

    for label, rn in [("פער", gap_r), ("% פער", pct_r)]:
        _s(ws.cell(rn, 1, label), bold=True, bg=C_GAP, h="right")
        for ci in range(2, 8):
            cl = get_column_letter(ci)
            c  = ws.cell(rn, ci)
            c.fill = _fill(C_GAP); c.border = _bdr()
            c.font = _f(bold=True); c.alignment = _al()
            if rn == gap_r:
                c.value = f"={cl}{act_r}-{cl}{bud_r}"
                c.number_format = KPI_FMTS[ci-2]
            else:
                c.value = f'=IF({cl}{bud_r}=0,"",({cl}{act_r}-{cl}{bud_r})/{cl}{bud_r})'
                c.number_format = "0.0%"

def _piece_table(ws, start_row, piece_daily, title="הכנסה שאינה שעתית"):
    """
    Write a 4-column side table for non-hourly days (יומי + תפוקה).
    Returns the row AFTER the table (for callers to continue writing below).
    If piece_daily is empty → no table written, returns start_row unchanged.

    Columns: תאריך | סוג | עובדים (or "אין נתון") | סכום (₪)
    """
    if not piece_daily:
        return start_row

    piece_sorted = sorted(piece_daily, key=_piece_sort_key)

    # Title bar (A:D)
    ws.merge_cells(f"A{start_row}:D{start_row}")
    c = ws.cell(start_row, 1, title)
    _s(c, bold=True, size=11, fg=C_WHITE, bg=C_PIECE_HDR, h="center")
    ws.row_dimensions[start_row].height = 20

    # Subtitle / note (merged A:D)
    note_r = start_row + 1
    ws.merge_cells(f"A{note_r}:D{note_r}")
    cn = ws.cell(note_r, 1, "לא נכלל בחישוב תעריף / שעות / הכנסה שעתית")
    _s(cn, bold=False, size=9, fg=C_BLACK, bg=C_PIECE, h="center", wrap=False)
    ws.row_dimensions[note_r].height = 14

    # Headers
    hdr_r = start_row + 2
    for ci, h in enumerate(["תאריך", "סוג", "עובדים", "סכום (₪)"], 1):
        _s(ws.cell(hdr_r, ci, h), bold=True, fg=C_WHITE, bg=C_PIECE_HDR)
    ws.row_dimensions[hdr_r].height = 16

    # Data rows
    for i, row in enumerate(piece_sorted):
        r  = hdr_r + 1 + i
        bg = C_PIECE
        _s(ws.cell(r, 1, row["date_label"]),              bg=bg, h="right")
        _s(ws.cell(r, 2, row.get("piece_type", "")),      bg=bg, h="center")
        if row["workers"] is None:
            _s(ws.cell(r, 3, "אין נתון"), bg=bg, h="center")
        else:
            _s(ws.cell(r, 3, int(row["workers"])), bg=bg, fmt="#,##0")
        _s(ws.cell(r, 4, row["payment"]), bg=bg, fmt="#,##0")

    last = hdr_r + len(piece_sorted)

    # Totals row
    tr = last + 1
    _s(ws.cell(tr, 1, 'סה"כ'), bold=True, bg=C_PIECE_HDR, fg=C_WHITE, h="right")
    _s(ws.cell(tr, 2, f"{len(piece_sorted)} ימים"),
       bold=True, bg=C_PIECE_HDR, fg=C_WHITE, h="center")
    # Workers total in col 3 — blank (may mix known + "אין נתון")
    _s(ws.cell(tr, 3, ""), bold=True, bg=C_PIECE_HDR, fg=C_WHITE, h="center")
    _s(ws.cell(tr, 4, sum(r["payment"] for r in piece_sorted)),
       bold=True, bg=C_PIECE_HDR, fg=C_WHITE, fmt="#,##0")

    return tr + 1

def write_monthly_sheet(ws, school, month_num, year, daily, bud, piece_daily=None):
    """One month tab: daily table + summary block + (optional) piece-rate side table."""
    ws.sheet_view.rightToLeft = True

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]; c.value = f"ניתוח פערים – {school} – {MONTH_HE[month_num]} {year}"
    _s(c, bold=True, size=13, fg=C_WHITE, bg=C_TITLE, h="center")
    ws.row_dimensions[1].height = 24

    # Daily header
    for ci, (h, f) in enumerate(zip(DAILY_HDRS, DAILY_FMTS), 1):
        c = ws.cell(2, ci, h)
        _s(c, bold=True, fg=C_WHITE, bg=C_HDR)
    ws.row_dimensions[2].height = 18

    # Daily rows
    for i, row in enumerate(daily or []):
        r  = 3 + i
        bg = C_ALT if i % 2 else None
        vals = [row["date_label"], int(row["workers"]), row["hours"],
                row["hours_per_worker"], row["payment"], row["rate"]]
        for ci, (v, fmt) in enumerate(zip(vals, DAILY_FMTS), 1):
            _s(ws.cell(r, ci, v), bg=bg, h="right" if ci == 1 else "center", fmt=fmt)

    last = 3 + len(daily or []) - 1

    # KPI header row
    sep = last + 3
    ws.merge_cells(f"A{sep}:G{sep}")
    c = ws.cell(sep, 1, "סיכום חודשי")
    _s(c, bold=True, fg=C_WHITE, bg=C_TITLE, h="center")
    ws.row_dimensions[sep].height = 18

    hdr_r = sep + 1
    _s(ws.cell(hdr_r, 1, ""), bg=C_HDR)
    for ci, (h, f) in enumerate(zip(KPI_HDRS, KPI_FMTS), 2):
        _s(ws.cell(hdr_r, ci, h), bold=True, fg=C_WHITE, bg=C_HDR)
    ws.row_dimensions[hdr_r].height = 16

    # Enable wrap on KPI header row so long headings fit
    for ci in range(1, 8):
        ws.cell(hdr_r, ci).alignment = _al("center", wrap=True)
    ws.row_dimensions[hdr_r].height = 32

    _summary_block(ws, sep + 2, actuals_kpis(daily), budget_kpis(bud))

    # Piece-rate side table — appears 2 rows after the summary block ends.
    # _summary_block uses rows act_r..act_r+4, so last used row = sep+6.
    if piece_daily:
        _piece_table(ws, sep + 8, piece_daily)

    # Column widths: max of (daily-table needs, summary-block needs)
    # A: date/label  B: workers/ימי-עבודה  C: סך-שעות/ממוצע-עובדים
    # D: שעות/עובד/ממוצע-שעות  E: הכנסה/שעות-עובד  F: תעריף/הכנסה  G: תעריף-summary
    cw(ws, {"A":13, "B":12, "C":18, "D":16, "E":14, "F":22, "G":24})
    ws.freeze_panes = "B1"   # lock leftmost column (date label stays visible on horizontal scroll)


def write_ytd_sheet(ws, school, year, through_month, monthly_daily, monthly_budgets,
                    monthly_piece=None):
    """
    YTD tab:
      Top section  – monthly summary table (one row per month, 4 key KPIs)
      Bottom section – full 6-KPI YTD totals vs budget
      Piece-rate side table – ALL piece days from Jan→through_month (flat, sorted)
    monthly_daily:   {month_num: [daily dicts] or []}
    monthly_budgets: {month_num: budget_dict or None}
    monthly_piece:   {month_num: [piece dicts] or []}  (optional)
    """
    ws.sheet_view.rightToLeft = True

    # Title — span full table width (11 cols: A through K)
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = f"סיכום מינואר עד {MONTH_HE[through_month]} {year} – {school}"
    _s(c, bold=True, size=13, fg=C_WHITE, bg=C_TITLE, h="center")
    ws.row_dimensions[1].height = 28

    # ── Monthly summary table ──────────────────────────────────────────────────
    # Cols: A=חודש | B-C=ימי עבודה | D-E=עובדים | F-G=שעות/עובד | H-I=הכנסה | J-K=תעריף
    GRP = [("ימי עבודה", 2), ("ממוצע עובדים/יום", 4),
           ("שעות/עובד", 6), ("הכנסה (₪)", 8), ("תעריף ממוצע בפועל (₪/ש)", 10)]
    ws.merge_cells("A2:A3"); _s(ws["A2"], bold=True, fg=C_WHITE, bg=C_HDR); ws["A2"].value = "חודש"
    for title, cs in GRP:
        ws.merge_cells(f"{get_column_letter(cs)}2:{get_column_letter(cs+1)}2")
        c2 = ws.cell(2, cs, title)
        _s(c2, bold=True, fg=C_WHITE, bg=C_HDR, wrap=True)
        for j, (sub, sbg) in enumerate([("ביצוע", C_ACT), ("תקציב", C_BUD)]):
            _s(ws.cell(3, cs+j, sub), bold=True, bg=sbg)
    ws.row_dimensions[2].height = 32; ws.row_dimensions[3].height = 16

    months = sorted(monthly_daily.keys())
    first_data = 4
    for i, m in enumerate(months):
        r    = first_data + i
        act  = actuals_kpis(monthly_daily[m])
        bud  = budget_kpis(monthly_budgets.get(m))
        ws.cell(r, 1, MONTH_HE[m]).alignment = _al("right"); ws.cell(r,1).border = _bdr()
        # cs: (act_idx, bud_idx, fmt)
        kpi_idx = {2:  (0, 0, "#,##0"),    # ימי עבודה
                   4:  (1, 1, "#,##0.0"),  # עובדים
                   6:  (3, 3, "#,##0.0"),  # שעות/עובד
                   8:  (4, 4, "#,##0"),    # הכנסה
                   10: (5, 5, "#,##0.00")} # תעריף
        for cs, (ai, bi, fmt) in kpi_idx.items():
            _s(ws.cell(r, cs,   act[ai]),                   bg=C_ACT, fmt=fmt)
            _s(ws.cell(r, cs+1, bud[bi] if bud else None),  bg=C_BUD, fmt=fmt)

    last_data = first_data + len(months) - 1

    # ── YTD totals section ─────────────────────────────────────────────────────
    all_daily = [d for m in months for d in (monthly_daily[m] or [])]
    bud_ytd   = ytd_budget(school, through_month)

    sep = last_data + 3
    # Bottom section title spans only the 7-col KPI block (A-G)
    ws.merge_cells(f"A{sep}:G{sep}")
    c = ws.cell(sep, 1, f"סיכום מצטבר מינואר עד {MONTH_HE[through_month]}")
    _s(c, bold=True, fg=C_WHITE, bg=C_TITLE, h="center")
    ws.row_dimensions[sep].height = 22

    hdr_r = sep + 1
    _s(ws.cell(hdr_r, 1, ""), bg=C_HDR)
    for ci, h in enumerate(KPI_HDRS, 2):
        _s(ws.cell(hdr_r, ci, h), bold=True, fg=C_WHITE, bg=C_HDR, wrap=True)
    ws.row_dimensions[hdr_r].height = 36   # tall row for wrapped KPI headers

    _summary_block(ws, sep + 2, actuals_kpis(all_daily), budget_kpis(bud_ytd))

    # ── Piece-rate side table (flat, all months) ──────────────────────────────
    # Place 2 rows below the YTD summary block's last row (sep+6).
    if monthly_piece:
        all_piece = [p for m in sorted(monthly_piece.keys())
                       for p in (monthly_piece[m] or [])]
        if all_piece:
            _piece_table(ws, sep + 8, all_piece,
                         title=f"הכנסה שאינה שעתית – מינואר עד {MONTH_HE[through_month]}")

    # Column widths: satisfy both the 11-col monthly table and the 7-col YTD block
    # Monthly table:  A=חודש | B-C=ימי-עבודה | D-E=עובדים | F-G=שעות/עובד | H-I=הכנסה | J-K=תעריף
    # YTD block:      A=label | B=ימי-עבודה | C=ממוצע-עובדים | D=ממוצע-שעות | E=שעות/עובד | F=הכנסה | G=תעריף
    cw(ws, {"A":16, "B":13, "C":16, "D":16, "E":14, "F":15, "G":14,
            "H":15, "I":15, "J":22, "K":22})
    ws.freeze_panes = "B1"   # lock month column


# ── Per-school annual workbook ─────────────────────────────────────────────────
def write_school_annual(school, months_daily, months_piece, through_month, year, path):
    """
    months_daily: {month_num: [daily dicts]}   (only months with hourly data)
    months_piece: {month_num: [piece dicts]}   (only months with piece data; may be {})
    """
    wb = Workbook(); wb.remove(wb.active)

    # YTD sheet first
    ws_ytd = wb.create_sheet("מינואר עד כה")
    monthly_budgets = {m: read_budget(school, m) for m in range(1, through_month+1)}
    write_ytd_sheet(ws_ytd, school, year, through_month,
                    months_daily, monthly_budgets, months_piece)

    # One sheet per month (only months that have data — hourly OR piece)
    all_months = set(months_daily.keys()) | set(months_piece.keys())
    for m in range(1, through_month + 1):
        if m not in all_months:
            continue
        ws = wb.create_sheet(MONTH_HE[m])
        write_monthly_sheet(ws, school, m, year,
                            months_daily.get(m, []),
                            read_budget(school, m),
                            months_piece.get(m, []))

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"    ✓  {os.path.basename(path)}")


# ── Network summary – one sheet per KPI ────────────────────────────────────────
# (kpi_title, kpi_index_in_actuals_kpis, number_format, is_summable_across_months)
NET_KPIS = [
    ("הכנסה (₪)",                4, "#,##0",    True),
    ("ימי עבודה",                 0, "#,##0",    True),
    ("ממוצע עובדים/יום",          1, "#,##0.0",  False),
    ("שעות/עובד",                 3, "#,##0.0",  False),
    ("תעריף ממוצע בפועל (₪/ש)",  5, "#,##0.00", False),
]

# Tab colours for the 5 KPI sheets (vivid but professional)
_TAB_COLORS = ["375623", "1F4E79", "7030A0", "C55A11", "833C00"]


def _weighted_total(pairs):
    """
    Weighted average across schools. Input: list of (value, weight) pairs.
    Returns Σ(v*w) / Σ(w), or None if total weight is 0 / no pairs.
    Used for network-level aggregation of rate-style KPIs, weighted by work_days —
    so that each (school, day) contributes equally to the network total.
    """
    if not pairs:
        return None
    total_w = sum(w for _, w in pairs if w)
    if total_w == 0:
        return None
    return sum(v * w for v, w in pairs if v is not None and w) / total_w


def _add_pct_cf(ws, range_str):
    """Conditional formatting: green if % > 0, red if % < 0."""
    gfill = PatternFill(start_color=C_GAP_POS, end_color=C_GAP_POS, fill_type="solid")
    rfill = PatternFill(start_color=C_GAP_NEG, end_color=C_GAP_NEG, fill_type="solid")
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=gfill, font=Font(name=FONT_NAME, bold=True, color=C_GAP_POS_F)))
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=rfill, font=Font(name=FONT_NAME, bold=True, color=C_GAP_NEG_F)))


def _write_net_kpi_sheet(wb, kpi_title, kpi_idx, kpi_fmt, summable,
                         schools_kpis, schools_ytd_kpis,
                         months, through_month, year, tab_color):
    """One professional sheet for a single KPI: schools×months (ביצוע/תקציב/% פער)."""
    safe_title = kpi_title.replace("/", "-").replace("\\", "-")[:31]
    ws = wb.create_sheet(safe_title)
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = tab_color

    schools   = sorted(schools_kpis)
    n_months  = len(months)

    def cs(mi): return 2 + mi * 3     # start col of month group (0-based month index)
    ytd_cs   = 2 + n_months * 3
    last_col = ytd_cs + 2

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(last_col)}1")
    c = ws["A1"]
    c.value = f"{kpi_title}  –  כלל הרשת {year}  (עד {MONTH_HE[through_month]})"
    _s(c, bold=True, size=13, fg=C_WHITE, bg=C_TITLE, h="center")
    ws.row_dimensions[1].height = 30

    # ── Rows 2-3: Month group headers ─────────────────────────────────────────
    ws.merge_cells("A2:A3")
    c = ws["A2"]; c.value = "בית ספר"
    _s(c, bold=True, fg=C_WHITE, bg=C_TITLE, h="center")

    for mi, m in enumerate(months):
        c0 = cs(mi)
        ws.merge_cells(f"{get_column_letter(c0)}2:{get_column_letter(c0+2)}2")
        _s(ws.cell(2, c0, MONTH_HE[m]), bold=True, fg=C_WHITE, bg=C_HDR)
        _s(ws.cell(3, c0,   "ביצוע"), bold=True, bg=C_ACT)
        _s(ws.cell(3, c0+1, "תקציב"), bold=True, bg=C_BUD)
        _s(ws.cell(3, c0+2, "% פער"), bold=True, bg=C_GAP)

    ws.merge_cells(f"{get_column_letter(ytd_cs)}2:{get_column_letter(ytd_cs+2)}2")
    _s(ws.cell(2, ytd_cs, "מינואר עד כה"), bold=True, fg=C_WHITE, bg=C_TITLE)
    _s(ws.cell(3, ytd_cs,   "ביצוע"), bold=True, bg=C_ACT)
    _s(ws.cell(3, ytd_cs+1, "תקציב"), bold=True, bg=C_BUD)
    _s(ws.cell(3, ytd_cs+2, "% פער"), bold=True, bg=C_GAP)
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 16

    # ── Data rows ─────────────────────────────────────────────────────────────
    fd = 4
    for si, school in enumerate(schools):
        r  = fd + si
        bg = C_ALT if si % 2 else None
        c  = ws.cell(r, 1, school)
        _s(c, h="right", bg=bg)
        c.font = _f(bold=True)

        for mi, m in enumerate(months):
            c0   = cs(mi)
            act  = schools_kpis[school].get(m)
            bkp  = budget_kpis(read_budget(school, m))
            av   = act[kpi_idx] if act else None
            bv   = bkp[kpi_idx] if bkp else None
            _s(ws.cell(r, c0,   av), bg=C_ACT, fmt=kpi_fmt)
            _s(ws.cell(r, c0+1, bv), bg=C_BUD, fmt=kpi_fmt)
            cl_a = get_column_letter(c0); cl_b = get_column_letter(c0+1)
            pc   = ws.cell(r, c0+2)
            pc.value        = f'=IF({cl_b}{r}=0,"",({cl_a}{r}-{cl_b}{r})/{cl_b}{r})'
            pc.fill         = _fill(C_GAP); pc.border = _bdr()
            pc.font         = _f(bold=True); pc.alignment = _al()
            pc.number_format = "0.0%"

        # YTD
        ytd_act = schools_ytd_kpis.get(school)
        ytd_bud = ytd_budget(school, through_month)
        ytd_bkp = budget_kpis(ytd_bud)
        av_y = ytd_act[kpi_idx] if ytd_act else None
        bv_y = ytd_bkp[kpi_idx] if ytd_bkp else None
        _s(ws.cell(r, ytd_cs,   av_y), bg=C_ACT, fmt=kpi_fmt)
        _s(ws.cell(r, ytd_cs+1, bv_y), bg=C_BUD, fmt=kpi_fmt)
        cl_a = get_column_letter(ytd_cs); cl_b = get_column_letter(ytd_cs+1)
        pc   = ws.cell(r, ytd_cs+2)
        pc.value        = f'=IF({cl_b}{r}=0,"",({cl_a}{r}-{cl_b}{r})/{cl_b}{r})'
        pc.fill         = _fill(C_GAP); pc.border = _bdr()
        pc.font         = _f(bold=True); pc.alignment = _al()
        pc.number_format = "0.0%"

    last_data = fd + len(schools) - 1

    # ── Totals row ────────────────────────────────────────────────────────────
    # Summable KPIs (הכנסה, ימי עבודה) → SUM Excel formula.
    # Rate-style KPIs (עובדים/יום, שעות/עובד, תעריף) → weighted average by
    #   each school's work_days, so each school-day contributes equally.
    tr = last_data + 1
    c  = ws.cell(tr, 1, 'סה"כ רשת')
    _s(c, bold=True, fg=C_WHITE, bg=C_TITLE, h="right")
    ws.row_dimensions[tr].height = 20

    def _style_total(cc, fmt=kpi_fmt):
        cc.fill          = _fill(C_TITLE); cc.border = _bdr()
        cc.font          = _f(bold=True, color=C_WHITE); cc.alignment = _al()
        cc.number_format = fmt

    for mi, m in enumerate(months):
        c0 = cs(mi)
        # Actual (j=0) and Budget (j=1)
        for j, data_getter in enumerate([
            lambda s: schools_kpis[s].get(m),
            lambda s: budget_kpis(read_budget(s, m)),
        ]):
            col = c0 + j
            cc  = ws.cell(tr, col)
            if summable:
                cl = get_column_letter(col)
                cc.value = f"=SUM({cl}{fd}:{cl}{last_data})"
            else:
                pairs = []
                for s in schools:
                    d = data_getter(s)
                    if d and d[kpi_idx] is not None and d[0]:
                        pairs.append((d[kpi_idx], d[0]))  # (value, work_days)
                cc.value = _weighted_total(pairs)
            _style_total(cc)
        # % פער (stays the same — references the cells above)
        cl_a = get_column_letter(c0); cl_b = get_column_letter(c0+1)
        pc   = ws.cell(tr, c0+2)
        pc.value = f'=IF({cl_b}{tr}=0,"",({cl_a}{tr}-{cl_b}{tr})/{cl_b}{tr})'
        _style_total(pc, fmt="0.0%")

    # YTD totals — same logic, but using YTD data (actual + budget)
    for j, data_getter in enumerate([
        lambda s: schools_ytd_kpis.get(s),
        lambda s: budget_kpis(ytd_budget(s, through_month)),
    ]):
        col = ytd_cs + j
        cc  = ws.cell(tr, col)
        if summable:
            cl = get_column_letter(col)
            cc.value = f"=SUM({cl}{fd}:{cl}{last_data})"
        else:
            pairs = []
            for s in schools:
                d = data_getter(s)
                if d and d[kpi_idx] is not None and d[0]:
                    pairs.append((d[kpi_idx], d[0]))
            cc.value = _weighted_total(pairs)
        _style_total(cc)
    cl_a = get_column_letter(ytd_cs); cl_b = get_column_letter(ytd_cs+1)
    pc   = ws.cell(tr, ytd_cs+2)
    pc.value = f'=IF({cl_b}{tr}=0,"",({cl_a}{tr}-{cl_b}{tr})/{cl_b}{tr})'
    _style_total(pc, fmt="0.0%")

    # ── Conditional formatting on all % פער columns ───────────────────────────
    for mi in range(n_months):
        cl = get_column_letter(cs(mi) + 2)
        _add_pct_cf(ws, f"{cl}{fd}:{cl}{last_data}")
    _add_pct_cf(ws, f"{get_column_letter(ytd_cs+2)}{fd}:{get_column_letter(ytd_cs+2)}{last_data}")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    for col in range(2, last_col + 1):
        offset = (col - 2) % 3
        ws.column_dimensions[get_column_letter(col)].width = 9 if offset == 2 else 15

    ws.freeze_panes = "B1"   # lock school column on horizontal scroll


def _write_net_piece_sheet(wb, schools_piece, months, through_month, year):
    """
    Network non-hourly income sheet: schools × months, actuals only (no budget).
    Aggregates both יומי and תפוקה into a single revenue figure per school/month.
    schools_piece: {school: {month_num: total_non_hourly_payment}}
    Creates sheet "הכנסה שאינה שעתית". If no school has any non-hourly income → skipped.
    """
    if not any(schools_piece.get(s) for s in schools_piece):
        return

    ws = wb.create_sheet("הכנסה שאינה שעתית")
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = C_PIECE_HDR

    schools  = sorted([s for s in schools_piece if schools_piece[s]])
    n_months = len(months)
    # Layout: A=school | B..=months | last=YTD
    ytd_col  = 2 + n_months
    last_col = ytd_col

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(last_col)}1")
    c = ws["A1"]
    c.value = f"הכנסה שאינה שעתית  –  כלל הרשת {year}  (עד {MONTH_HE[through_month]})"
    _s(c, bold=True, size=13, fg=C_WHITE, bg=C_PIECE_HDR, h="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: explanatory subtitle ───────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(last_col)}2")
    c = ws["A2"]
    c.value = "תשלומים יומיים ולפי תפוקה – לא נכללים ב-KPI השעתיים של הרשת"
    _s(c, bold=False, size=9, fg=C_BLACK, bg=C_PIECE, h="center")
    ws.row_dimensions[2].height = 14

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    _s(ws.cell(3, 1, "בית ספר"), bold=True, fg=C_WHITE, bg=C_TITLE, h="center")
    for mi, m in enumerate(months):
        _s(ws.cell(3, 2 + mi, MONTH_HE[m]), bold=True, fg=C_WHITE, bg=C_HDR)
    _s(ws.cell(3, ytd_col, "מינואר עד כה"), bold=True, fg=C_WHITE, bg=C_TITLE)
    ws.row_dimensions[3].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    fd = 4
    for si, school in enumerate(schools):
        r  = fd + si
        bg = C_ALT if si % 2 else None
        c  = ws.cell(r, 1, school)
        _s(c, h="right", bg=bg)
        c.font = _f(bold=True)

        ytd_total = 0
        for mi, m in enumerate(months):
            v = schools_piece[school].get(m, 0)
            _s(ws.cell(r, 2 + mi, v if v else None), bg=C_PIECE, fmt="#,##0")
            ytd_total += v
        _s(ws.cell(r, ytd_col, ytd_total if ytd_total else None),
           bg=C_ACT, fmt="#,##0", bold=True)

    last_data = fd + len(schools) - 1

    # ── Totals row ────────────────────────────────────────────────────────────
    tr = last_data + 1
    _s(ws.cell(tr, 1, 'סה"כ רשת'), bold=True, fg=C_WHITE, bg=C_TITLE, h="right")
    ws.row_dimensions[tr].height = 20

    for mi in range(n_months):
        col = 2 + mi; cl = get_column_letter(col)
        cc  = ws.cell(tr, col)
        cc.value         = f"=SUM({cl}{fd}:{cl}{last_data})"
        cc.fill          = _fill(C_TITLE); cc.border = _bdr()
        cc.font          = _f(bold=True, color=C_WHITE); cc.alignment = _al()
        cc.number_format = "#,##0"
    cl = get_column_letter(ytd_col)
    cc = ws.cell(tr, ytd_col)
    cc.value         = f"=SUM({cl}{fd}:{cl}{last_data})"
    cc.fill          = _fill(C_TITLE); cc.border = _bdr()
    cc.font          = _f(bold=True, color=C_WHITE); cc.alignment = _al()
    cc.number_format = "#,##0"

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    for col in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = "B1"   # lock school column on horizontal scroll


def _write_net_dashboard_sheet(wb, schools_kpis, schools_ytd_kpis, schools_piece,
                                through_month, year):
    """
    'תצוגת על' dashboard: FIRST sheet in the network workbook.
    One row per school, YTD values for ALL 5 KPIs (ביצוע/תקציב/% פער) + non-hourly income,
    with conditional formatting on gap columns and a network totals row.
    """
    ws = wb.create_sheet("תצוגת על", 0)   # index 0 → first sheet
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = C_TITLE

    schools = sorted(schools_ytd_kpis)
    n_kpis  = len(NET_KPIS)   # 5

    # Column layout:
    #   A      = school
    #   B..    = 5 KPI groups × 3 cols (ביצוע/תקציב/%) = 15 cols
    #   last   = הכנסה לא שעתית (single col, actual only)
    def kpi_cs(ki): return 2 + ki * 3
    piece_col = 2 + n_kpis * 3
    last_col  = piece_col

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(last_col)}1")
    c = ws["A1"]
    c.value = f"תצוגת על – כלל הרשת {year}   (YTD ינואר–{MONTH_HE[through_month]})"
    _s(c, bold=True, size=14, fg=C_WHITE, bg=C_TITLE, h="center")
    ws.row_dimensions[1].height = 32

    # ── Rows 2-3: group headers ───────────────────────────────────────────────
    ws.merge_cells("A2:A3")
    c = ws["A2"]; c.value = "בית ספר"
    _s(c, bold=True, fg=C_WHITE, bg=C_TITLE, h="center")

    for ki, (kpi_title, kpi_idx, kpi_fmt, summable) in enumerate(NET_KPIS):
        cs = kpi_cs(ki)
        ws.merge_cells(f"{get_column_letter(cs)}2:{get_column_letter(cs+2)}2")
        _s(ws.cell(2, cs, kpi_title), bold=True, fg=C_WHITE, bg=C_HDR, wrap=True)
        _s(ws.cell(3, cs,   "ביצוע"), bold=True, bg=C_ACT)
        _s(ws.cell(3, cs+1, "תקציב"), bold=True, bg=C_BUD)
        _s(ws.cell(3, cs+2, "% פער"), bold=True, bg=C_GAP)

    # Non-hourly column — single cell spanning rows 2-3
    ws.merge_cells(f"{get_column_letter(piece_col)}2:{get_column_letter(piece_col)}3")
    _s(ws.cell(2, piece_col, "הכנסה לא שעתית (YTD)"),
       bold=True, fg=C_WHITE, bg=C_PIECE_HDR, wrap=True)

    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────────
    fd = 4
    for si, school in enumerate(schools):
        r  = fd + si
        bg = C_ALT if si % 2 else None
        c  = ws.cell(r, 1, school)
        _s(c, h="right", bg=bg); c.font = _f(bold=True)

        ytd_act = schools_ytd_kpis.get(school)
        ytd_bud = budget_kpis(ytd_budget(school, through_month))

        for ki, (kpi_title, kpi_idx, kpi_fmt, summable) in enumerate(NET_KPIS):
            cs = kpi_cs(ki)
            av = ytd_act[kpi_idx] if ytd_act else None
            bv = ytd_bud[kpi_idx] if ytd_bud else None
            _s(ws.cell(r, cs,   av), bg=C_ACT, fmt=kpi_fmt)
            _s(ws.cell(r, cs+1, bv), bg=C_BUD, fmt=kpi_fmt)
            cl_a = get_column_letter(cs); cl_b = get_column_letter(cs+1)
            pc   = ws.cell(r, cs+2)
            pc.value        = f'=IF({cl_b}{r}=0,"",({cl_a}{r}-{cl_b}{r})/{cl_b}{r})'
            pc.fill         = _fill(C_GAP); pc.border = _bdr()
            pc.font         = _f(bold=True); pc.alignment = _al()
            pc.number_format = "0.0%"

        # Non-hourly YTD total for this school
        piece_total = sum(schools_piece.get(school, {}).values())
        _s(ws.cell(r, piece_col, piece_total if piece_total else None),
           bg=C_PIECE, fmt="#,##0")

    last_data = fd + len(schools) - 1

    # ── Totals row ────────────────────────────────────────────────────────────
    # Same philosophy as the per-KPI sheets:
    #   summable → SUM; rate-style → weighted avg by each school's YTD work_days.
    tr = last_data + 1
    _s(ws.cell(tr, 1, 'סה"כ רשת'), bold=True, fg=C_WHITE, bg=C_TITLE, h="right")
    ws.row_dimensions[tr].height = 22

    def _style_total(cc, fmt):
        cc.fill          = _fill(C_TITLE); cc.border = _bdr()
        cc.font          = _f(bold=True, color=C_WHITE); cc.alignment = _al()
        cc.number_format = fmt

    for ki, (kpi_title, kpi_idx, kpi_fmt, summable) in enumerate(NET_KPIS):
        cs_k = kpi_cs(ki)
        for j, data_getter in enumerate([
            lambda s: schools_ytd_kpis.get(s),
            lambda s: budget_kpis(ytd_budget(s, through_month)),
        ]):
            col = cs_k + j
            cc  = ws.cell(tr, col)
            if summable:
                cl = get_column_letter(col)
                cc.value = f"=SUM({cl}{fd}:{cl}{last_data})"
            else:
                pairs = []
                for s in schools:
                    d = data_getter(s)
                    if d and d[kpi_idx] is not None and d[0]:
                        pairs.append((d[kpi_idx], d[0]))
                cc.value = _weighted_total(pairs)
            _style_total(cc, kpi_fmt)
        cl_a = get_column_letter(cs_k); cl_b = get_column_letter(cs_k+1)
        pc   = ws.cell(tr, cs_k+2)
        pc.value = f'=IF({cl_b}{tr}=0,"",({cl_a}{tr}-{cl_b}{tr})/{cl_b}{tr})'
        _style_total(pc, "0.0%")

    # Non-hourly total — plain SUM (summable revenue)
    cl = get_column_letter(piece_col)
    cc = ws.cell(tr, piece_col)
    cc.value = f"=SUM({cl}{fd}:{cl}{last_data})"
    _style_total(cc, "#,##0")

    # ── Conditional formatting on all % פער columns ───────────────────────────
    for ki in range(n_kpis):
        cl = get_column_letter(kpi_cs(ki) + 2)
        _add_pct_cf(ws, f"{cl}{fd}:{cl}{last_data}")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 20
    for ki in range(n_kpis):
        cs = kpi_cs(ki)
        ws.column_dimensions[get_column_letter(cs)].width   = 13
        ws.column_dimensions[get_column_letter(cs+1)].width = 13
        ws.column_dimensions[get_column_letter(cs+2)].width = 9
    ws.column_dimensions[get_column_letter(piece_col)].width = 18

    ws.freeze_panes = "B1"   # lock school column on horizontal scroll


def write_network_summary(schools_kpis, schools_ytd_kpis, schools_piece,
                          through_month, year, path):
    """
    Create a professional network summary workbook with:
      1. 'תצוגת על' — FIRST sheet, all KPIs at a glance (YTD, all schools).
      2. One sheet per KPI — drill-down schools × months (5 sheets).
      3. 'הכנסה שאינה שעתית' — only if any school has non-hourly income.
    schools_kpis:     {school: {month_num: actuals_kpis list}}
    schools_ytd_kpis: {school: actuals_kpis list (YTD aggregate)}
    schools_piece:    {school: {month_num: piece payment total}}
    """
    wb = Workbook()
    wb.remove(wb.active)
    months = sorted({m for sk in schools_kpis.values() for m in sk}
                    | {m for sk in schools_piece.values() for m in sk})

    # Dashboard first (becomes sheet index 0)
    _write_net_dashboard_sheet(wb, schools_kpis, schools_ytd_kpis, schools_piece,
                               through_month, year)

    for i, (kpi_title, kpi_idx, kpi_fmt, summable) in enumerate(NET_KPIS):
        _write_net_kpi_sheet(
            wb, kpi_title, kpi_idx, kpi_fmt, summable,
            schools_kpis, schools_ytd_kpis, months, through_month, year,
            _TAB_COLORS[i % len(_TAB_COLORS)]
        )

    # Piece-rate network sheet (only if relevant)
    _write_net_piece_sheet(wb, schools_piece, months, through_month, year)

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"    ✓  {os.path.basename(path)}")


# ── Main ───────────────────────────────────────────────────────────────────────
def run(year, month):
    print(f"\n{'='*60}")
    print(f"  ניתוח שנתי עד: {MONTH_HE[month]} {year}")
    print(f"{'='*60}\n")

    # Load data for all months Jan → month
    # all_data[m][school]       = list of daily dicts   (hourly)
    # all_data_piece[m][school] = list of piece dicts   (piece-rate, aggregated by date)
    all_data       = {}
    all_data_piece = {}
    all_warnings   = []
    for m in range(1, month + 1):
        src = monthly_path(year, m)
        if not os.path.exists(src):
            print(f"  ⚠  חסר קובץ: {os.path.basename(src)} – דולג")
            continue
        print(f"  קורא: {os.path.basename(src)}")
        wb  = load_workbook(src, data_only=True)
        all_data[m]       = {}
        all_data_piece[m] = {}
        for school in school_sheets(wb):
            rate = read_budget_rate(school)
            hourly_rows, piece_rows, warns = read_school_ws(wb[school], school_rate=rate)
            for w in warns:
                all_warnings.append(f"{MONTH_HE[m]} | {school} | {w}")
            if hourly_rows:
                all_data[m][school] = aggregate_by_date(hourly_rows)
            if piece_rows:
                all_data_piece[m][school] = aggregate_piece_by_date(piece_rows)
            n_h = len(all_data[m].get(school, []))
            n_p = len(all_data_piece[m].get(school, []))
            if n_h or n_p:
                msg = f"    {school}: {n_h} ימי עבודה"
                if n_p:
                    msg += f"  +{n_p} ימים לא-שעתיים"
                if warns:
                    msg += f"  ⚠ {len(warns)} שורות חסרות"
                print(msg)

    if not all_data and not all_data_piece:
        print("\n❌ לא נמצאו נתונים"); sys.exit(1)

    # Collect all schools (union from hourly + piece)
    all_schools = sorted({s for m_d in all_data.values()       for s in m_d}
                        | {s for m_d in all_data_piece.values() for s in m_d})
    print(f"\n  בתי ספר: {', '.join(all_schools)}\n")

    out_folder = os.path.join(OUT_DIR, f"{year}-{month:02d}-{MONTH_HE[month]}")

    # Build schools_kpis + YTD kpis for network summary (hourly only — unchanged)
    schools_kpis     = {}
    schools_ytd_kpis = {}
    # schools_piece: {school: {month: total_piece_payment}}
    schools_piece    = {}

    for school in all_schools:
        print(f"  ► {school}")
        months_daily = {m: all_data[m].get(school, [])
                        for m in all_data
                        if school in all_data[m]}
        months_piece = {m: all_data_piece[m].get(school, [])
                        for m in all_data_piece
                        if school in all_data_piece[m]}

        schools_kpis[school] = {m: actuals_kpis(daily)
                                 for m, daily in months_daily.items()}

        # Proper YTD kpis (weighted rate, true averages) from raw daily data
        all_school_daily = [d for daily in months_daily.values() for d in daily]
        schools_ytd_kpis[school] = actuals_kpis(all_school_daily)

        # Piece totals per month (for network sheet)
        schools_piece[school] = {m: sum(p["payment"] for p in rows)
                                 for m, rows in months_piece.items()
                                 if rows}

        yy = str(year)[-2:]
        out_file = os.path.join(out_folder, f"ניתוח פערים {school} {month:02d}-{yy}.xlsx")
        write_school_annual(school, months_daily, months_piece,
                            month, year, out_file)

    # Network summary
    print(f"\n  ► סיכום רשת")
    write_network_summary(schools_kpis, schools_ytd_kpis, schools_piece,
                          month, year,
                          os.path.join(out_folder, f"ניתוח פערים סיכום רשת {month:02d}-{str(year)[-2:]}.xlsx"))

    # ── Write warnings log ────────────────────────────────────────────────────
    log_path = os.path.join(out_folder, f"לוג בעיות {month:02d}-{str(year)[-2:]}.txt")
    os.makedirs(out_folder, exist_ok=True)
    with open(log_path, "w", encoding="utf-8") as f:
        f.write(f"לוג בעיות נתונים – {MONTH_HE[month]} {year}\n")
        f.write("=" * 50 + "\n\n")
        if all_warnings:
            f.write(f"נמצאו {len(all_warnings)} שורות עם בעיות:\n\n")
            for w in all_warnings:
                f.write(w + "\n")
            f.write("\n⚠ יש לתקן את הנתונים בקובץ המקור ולהריץ מחדש.\n")
            f.write("ℹ  תזכורת: שורות לא-שעתיות (עמודה E מכילה 'תפוקה'/'יום') אינן מופיעות כאן –\n")
            f.write("   הן מטופלות בנפרד ומוצגות בטבלת 'הכנסה שאינה שעתית' בכל חוברת.\n")
        else:
            f.write("✅ לא נמצאו בעיות — כל הנתונים תקינים.\n")

    if all_warnings:
        print(f"\n⚠  נמצאו {len(all_warnings)} שורות עם שעות חסרות → {os.path.basename(log_path)}")
    else:
        print(f"\n✅ כל הנתונים תקינים")
    print(f"✅ הסתיים!\n   {out_folder}\n")


if __name__ == "__main__":
    if len(sys.argv) == 3:
        y, m = int(sys.argv[1]), int(sys.argv[2])
    elif len(sys.argv) == 1:
        now = datetime.now(); y, m = now.year, now.month
    else:
        print("Usage: python analyze_monthly.py [YYYY] [MM]"); sys.exit(1)
    if not 1 <= m <= 12:
        print(f"❌ חודש לא תקין: {m}"); sys.exit(1)
    run(y, m)
