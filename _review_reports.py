#!/usr/bin/env python3
"""Scan reports for noteworthy gaps before sending to management."""
import os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

REPORT_DIR = r"C:\Users\דרורנדל\HASHOMER HACHADASH\כספים - General\כספים - תכנון ובקרה\אדם ואדמה\2026\בקרות חודשיות\הכנסות מחקלאים\ניתוח פערים\דוחות\2026-03-מרץ"

SCHOOLS = ["אשכול", "גולן", "גליל עליון", "חצבה", "ניצנה", "ניצנים", "תבור"]

# ── 1. Network summary YTD ────────────────────────────────────────────────────
print("="*70)
print("YTD מינואר עד מרץ — סיכום רשת")
print("="*70)
net = os.path.join(REPORT_DIR, "ניתוח פערים סיכום רשת 03-26.xlsx")
wb = load_workbook(net, data_only=True)
ws = wb["תצוגת על"]

print(f"\n[תצוגת על] dims: {ws.max_row}x{ws.max_column}")
for r in range(1, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    if any(v is not None for v in row):
        print(f"  R{r}: {row}")

# ── 2. Per-school: monthly KPIs and deviations ─────────────────────────────────
print("\n" + "="*70)
print("ביצוע vs. תקציב — לכל בית ספר (YTD)")
print("="*70)

KPI_NAMES = ["ימי עבודה", "ממוצע עובדים/יום", "ממוצע שעות/יום",
             "שעות/עובד", "הכנסה (₪)", "תעריף ממוצע (₪/ש)"]

for school in SCHOOLS:
    fp = os.path.join(REPORT_DIR, f"ניתוח פערים {school} 03-26.xlsx")
    if not os.path.exists(fp):
        print(f"\n--- {school} --- MISSING"); continue
    wbs = load_workbook(fp, data_only=True)
    if "מינואר עד כה" not in wbs.sheetnames:
        print(f"\n--- {school} --- no YTD sheet"); continue
    yws = wbs["מינואר עד כה"]
    # Find YTD block: row labelled 'ביצוע' under 'סיכום מצטבר'
    print(f"\n--- {school} ---")
    # Print monthly table (top section)
    for r in range(1, min(8, yws.max_row + 1)):
        row = [yws.cell(r, c).value for c in range(1, 12)]
        if any(v is not None for v in row):
            cleaned = [round(v, 2) if isinstance(v, float) else v for v in row]
            print(f"  R{r}: {cleaned}")
    # Find and print YTD summary block
    for r in range(8, yws.max_row + 1):
        a = yws.cell(r, 1).value
        if a in ("ביצוע", "תקציב", "פער", "% פער"):
            row = [yws.cell(r, c).value for c in range(1, 8)]
            cleaned = [round(v, 2) if isinstance(v, float) else v for v in row]
            print(f"  R{r}: {cleaned}")
