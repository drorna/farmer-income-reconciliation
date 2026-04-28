#!/usr/bin/env python3
"""Inspect ניצנים March rows around 16-17."""
import os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

src = r"C:\Users\דרורנדל\HASHOMER HACHADASH\כספים - General\כספים - תכנון ובקרה\אדם ואדמה\2026\בקרות חודשיות\הכנסות מחקלאים\ניתוח פערים\נתונים מרוכזים\מרכז חשבוניות חקלאים מרץ 26.xlsx"
wb = load_workbook(src, data_only=True)
print("Sheets:", wb.sheetnames)
ws = wb["ניצנים"] if "ניצנים" in wb.sheetnames else None
if ws is None:
    print("ניצנים sheet not found"); sys.exit()

print(f"\nניצנים – {ws.max_row} rows x {ws.max_column} cols")
print("Showing rows where date contains 16 or 17:\n")
for r in range(1, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, min(7, ws.max_column + 1))]
    a = row[0]
    if a is None: continue
    s = str(a)
    if "16" in s or "17" in s or r <= 3:
        print(f"  R{r}: {row}")
